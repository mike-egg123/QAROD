import os
import random

import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from openpyxl.styles import Alignment
from pyod.models.copod import COPOD
from pyod.models.iforest import IForest
from pyod.models.knn import KNN
from pyod.models.lof import LOF
from pyod.models.suod import SUOD
from pyod.utils import evaluate_print
from sklearn.cluster import KMeans
from scipy import stats
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn import manifold
import banpei
from sklearn.decomposition import PCA, KernelPCA
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# 读取文件
def read_data(filename):
    df = pd.read_csv(filename)
    return df


# 提取qar数据表格中的所有参数
def extract_params(data_frame):
    cols = data_frame.columns.tolist()
    return cols


# 筛选出值为数值且有变化的参数
def valid_params(params, data_frame):
    v_params = []
    # 筛选出值为数值的参数
    for param in params:
        data = data_frame[param][3:]
        flag = False
        for d in data:
            try:
                float(d)
            except:
                flag = True
        if not flag:
            v_params.append(param)
    changed_v_params = []
    for param in v_params:
        data = data_frame[param].values.tolist()[3:]
        data = [float(t) for t in data]  # 转成浮点数
        # 判断该参数是否全为nan
        flag = False
        for d in data:
            if d == d:
                flag = True
                break
        if not flag:
            continue
        # 剔除不变化的参数
        first = data[0]
        for d in data:
            if d != first:
                changed_v_params.append(param)
                break
    return changed_v_params


# 滑动窗口函数
# values    所有数据
# size      数据长度
# step
def sliding_windows(values, size, step):
    first = sum(values[0:size])
    new_list = []
    new_list.append(first / size)
    for i in range(1, len(values) - size + 1, step):
        # 以步长为step滑动窗口，step越大压缩率越高
        for j in range(1, step + 1):
            first -= values[i - j]  # 滑出
            first += values[i + size - j]  # 滑入
        new_list.append(first / size)  # 均值
    return new_list


# 动态窗口均值差，需要传入每个点所处的飞行阶段flight_phases以及有多少处在该飞行阶段的时间点
def dynamic_window_mean(values, flight_phases, flight_phases_num, freq):
    size = int(len(flight_phases * freq) * 0.2)  # 还是要用边界值扩充窗口，但这个时候要将窗口扩充得尽可能大
    # 用边界值扩充窗口
    for i in range(0, size // 2):
        values.insert(0, values[0])
    for i in range(0, size // 2):
        values.append(values[-1])
    window_sum = sum(values[0:size])
    means = []
    # 计算每个点与以其为中心的窗口的均值的差距
    for i in range(size // 2, len(values) - size // 2):
        means.append(abs(values[i] - window_sum / size))
        if i == len(values) - size // 2 - 1:
            break
        actual_i = (i - size // 2) * freq
        # print(actual_i)
        diff = size // 2
        # 根据该点所处的飞行阶段来调整窗口大小，系数根据经验给出
        if flight_phases[actual_i] == 'CRUISE':
            diff = flight_phases_num['CRUISE'] * 0.2
        elif flight_phases[actual_i] == 'DESCENT':
            diff = flight_phases_num['DESCENT'] * 0.2
        elif flight_phases[actual_i] == 'CLIMB':
            diff = flight_phases_num['CLIMB'] * 0.1
        elif flight_phases[actual_i] == 'TAXI OUT':
            diff = flight_phases_num['TAXI OUT'] * 0.2
        elif flight_phases[actual_i] == 'TAXI IN':
            diff = flight_phases_num['TAXI IN'] * 0.2
        elif flight_phases[actual_i] == 'FINAL':
            diff = flight_phases_num['FINAL'] * 0.1
        elif flight_phases[actual_i] == 'APPROACH':
            diff = flight_phases_num['APPROACH'] * 0.05
        elif flight_phases[actual_i] == 'INI. CLIMB':
            diff = flight_phases_num['INI. CLIMB'] * 0.05
        elif flight_phases[actual_i] == 'TAKE OFF':
            diff = flight_phases_num['TAKE OFF'] * 0.05
        elif flight_phases[actual_i] == 'LANDING':
            diff = flight_phases_num['LANDING'] * 0.05
        elif flight_phases[actual_i] == 'ENG. STOP':
            diff = flight_phases_num['ENG. STOP'] * 0.2
        elif flight_phases[actual_i] == '2 SEGMENT':
            diff = flight_phases_num['2 SEGMENT'] * 0.05
        # window_sum -= values[0 if i - int(diff) < 0 else i - int(diff)]
        # window_sum += values[len(values) - 1 if i + int(diff) + 1 >= len(values) else i + int(diff) + 1]
        window_sum -= values[i - int(diff)]
        window_sum += values[i + int(diff) + 1]
    return means

# 静态窗口均值差
def static_window_mean(values, size):
    # 用边界值扩充窗口
    for i in range(0, size // 2):
        values.insert(0, values[0])
    for i in range(0, size // 2):
        values.append(values[-1])
    window_sum = sum(values[0:size])
    means = []
    # print(len(flight_phases))
    # 计算每个点与以其为中心的窗口的均值的差距
    for i in range(size // 2, len(values) - size // 2):
        means.append(abs(values[i] - window_sum / size))
        if i == len(values) - size // 2 - 1:
            break
        diff = size // 2
        window_sum -= values[i - int(diff)]
        window_sum += values[i + int(diff) + 1]
    return means



# 左右最大斜率
def point_slope(values):
    # 扩充边界值
    values.insert(0, values[0])
    values.append(values[-1])
    slopes = []
    for i in range(1, len(values) - 1):
        slopes.append(max(abs(values[i] - values[i - 1]), abs(values[i] - values[i + 1])))  # 取左斜率和右斜率中较大的值
    return slopes


# 肘部法寻找最佳的k
def elbow_best_k(values):
    sse = []
    # 计算k从2到10的sse
    for k in range(2, 10):
        km = KMeans(n_clusters=k).fit(np.array(values).reshape(-1, 1))
        sse.append(km.inertia_)
    # 计算k从2到10的sse的斜率，并对这串斜率再做一次K-Means聚类，指定聚成两类，此时处于两类交接处的k即为我们要找的“肘部”
    diff = []
    for i in range(1, len(sse)):
        diff.append(sse[i - 1] - sse[i])
    labels = KMeans(n_clusters=2).fit_predict(np.array(diff).reshape(-1, 1))
    for i in range(1, len(labels)):
        if labels[i] != labels[i - 1]:
            return i + 2, sse
    return 3, sse  # 给个默认情况，k=3

# 肘部法寻找最佳的k
def elbow_best_k_2d(values):
    sse = []
    # 计算k从2到10的sse
    for k in range(2, 10):
        km = KMeans(n_clusters=k).fit(values)
        sse.append(km.inertia_)
    # 计算k从2到10的sse的斜率，并对这串斜率再做一次K-Means聚类，指定聚成两类，此时处于两类交接处的k即为我们要找的“肘部”
    diff = []
    for i in range(1, len(sse)):
        diff.append(sse[i - 1] - sse[i])
    labels = KMeans(n_clusters=2).fit_predict(np.array(diff).reshape(-1, 1))
    for i in range(1, len(labels)):
        if labels[i] != labels[i - 1]:
            return i + 2, sse
    return 3, sse  # 给个默认情况，k=3


# 相关分析
# 计算所有可检测的参数之间的相关系数，以及p值
def get_corrcoef(params, df):
    print("正在进行参数间相关性分析，请稍后")
    # 将所有可检测的参数对应的数据放入lst中
    lst = []
    for param in params:
        data = df[param].values.tolist()[3:]
        data = [float(t) for t in data]
        # 补齐部分nan,解决频率不一致问题
        for i in range(0, len(data)):
            # 向后寻找非nan数来补齐
            if data[i] != data[i]:
                for j in range(i + 1, len(data)):
                    if data[j] == data[j]:
                        data[i] = data[j]
                        break
            # 向前寻找非nan数来补齐
            if data[i] != data[i]:
                for j in range(i - 1, -1, -1):
                    if data[j] == data[j]:
                        data[i] = data[j]
                        break
        lst.append(data)
    data_frame = pd.DataFrame(lst, index=params).T
    df_corr = pd.DataFrame()  # 存储相关系数矩阵
    df_p = pd.DataFrame()  # 存储p值矩阵
    for x in data_frame.columns:
        for y in data_frame.columns:
            corr = stats.pearsonr(data_frame[x], data_frame[y])
            df_corr.loc[x, y] = corr[0]
            df_p.loc[x, y] = corr[1]
    return df_corr, df_p

#  找出对应的参数最相关的五个参数
def get_most_related_params(corr, p, param):
    corrcoefs = abs(corr[param])  # 取绝对值，避免正负相关的影响
    corrcoefs = corrcoefs.sort_values(ascending=False)  # 降序排列
    p_values = p[param]
    most_related_params = []
    for i in range(0, len(corrcoefs)):
        # 避开自己
        if corrcoefs.index[i] == param:
            continue
        # 取四个最相关的参数
        if len(most_related_params) >= 4:
            break
        # 要0.5以上才有较强的相关性，0.05以下才有较显著的相关性
        if corrcoefs.iloc[i] > 0.5:
            if p_values[corrcoefs.index[i]] < 0.05:
                most_related_params.append(corrcoefs.index[i])
        else:
            break
    print(str(param), "的五个相关参数为：", most_related_params)
    return most_related_params


# 结合单参数分析的异常标记数组，进行多参数多元回归分析
def multi_regression(df, param, most_related_params, single_anomaly_points):
    # print(most_related_params)
    if len(most_related_params) == 0:
        return single_anomaly_points, []
    if len(single_anomaly_points) == 0:
        return [], []
    data = []
    # 标签值，即被选中的参数
    Y = df[param].values.tolist()[3:]
    Y = [float(t) for t in Y]
    # 补齐部分nan,解决频率不一致问题
    for i in range(0, len(Y)):
        # 向后寻找非nan数来补齐
        if Y[i] != Y[i]:
            for j in range(i + 1, len(Y)):
                if Y[j] == Y[j]:
                    Y[i] = Y[j]
                    break
        # 向前寻找非nan数来补齐
        if Y[i] != Y[i]:
            for j in range(i - 1, -1, -1):
                if Y[j] == Y[j]:
                    Y[i] = Y[j]
                    break
    Y_normal = []
    Y_abnormal = []
    # 分出正常和异常类
    for i in range(0, len(Y)):
        if i in single_anomaly_points:
            Y_abnormal.append(Y[i])
        else:
            Y_normal.append(Y[i])
    # print(len(Y_normal))
    # 特征值，即与被选中的参数相关性最强的5个参数
    xs_abnormal = []
    for related_param in most_related_params:
        x = df[related_param].values.tolist()[3:]
        x = [float(t) for t in x]
        # 补齐部分nan,解决频率不一致问题
        for i in range(0, len(x)):
            # 向后寻找非nan数来补齐
            if x[i] != x[i]:
                for j in range(i + 1, len(x)):
                    if x[j] == x[j]:
                        x[i] = x[j]
                        break
            # 向前寻找非nan数来补齐
            if x[i] != x[i]:
                for j in range(i - 1, -1, -1):
                    if x[j] == x[j]:
                        x[i] = x[j]
                        break
        x_normal = []
        x_abnormal = []
        for i in range(0, len(x)):
            if i in single_anomaly_points:
                x_abnormal.append(x[i])
            else:
                x_normal.append(x[i])
        # print(len(x_normal))
        data.append(x_normal)
        xs_abnormal.append(x_abnormal)
    # print(data)
    # print(Y_normal)
    # 划分训练集和阈值生成集8：2
    X_train, X_test, Y_train, Y_test = train_test_split(np.array(data).T, Y_normal, train_size=0.8)
    # 多元线性回归
    model = LinearRegression()
    model.fit(X_train, Y_train)
    a = model.intercept_  # 截距
    b = model.coef_  # 系数
    # print(a, "\n", b)
    score = model.score(X_test, Y_test)
    # print(score)
    Y_pred = model.predict(X_test)  # 预测值
    # print(len(Y_test), " ", len(Y_pred))
    # 生成平均阈值
    e = 0
    for i in range(0, len(Y_test)):
        e += abs(Y_pred[i] - Y_test[i])
    e /= len(Y_test)
    # print(e)
    Y_abnormal_pred = model.predict(np.array(xs_abnormal).T)
    # print(len(Y_abnormal_pred), len(Y_abnormal), len(single_anomaly_points))
    # 区分出两种异常：type1为传感器跳变异常，type2为误操作异常
    type1 = []
    type2 = []
    for i in range(0, len(Y_abnormal)):
        e_ = abs(Y_abnormal[i] - Y_abnormal_pred[i])
        if e_ > e:
            type1.append(single_anomaly_points[i])
        else:
            type2.append(single_anomaly_points[i])
    return type1, type2

# 返回数字的正负号
def sig(x):
    if x > 0:
        return 1
    elif x == 0:
        return 0
    else:
        return -1

# 以时间的角度来考虑异常原因
def time_analyse_cause(type1, type2, data):
    n_type1 = []
    n_type2 = []
    for i in range(1, len(data) - 1):
        if i in type2:
            n_type2.append(i)
            flag_left = sig(data[i] - data[i - 1])
            flag_right = sig(data[i + 1] - data[i])
            # 左侧
            for j in range(i - 1, 0, -1):
                if sig(data[j] - data[j - 1] ) != flag_left:  # 出现趋势的改变，退出染色
                    break
                if j in type2:  # 搜到同类，退出染色
                    break
                if j in type1:  # 搜到另一类，染色
                    n_type2.append(j)
            # 右侧
            for j in range(i + 1, len(data) - 1):
                if sig(data[j + 1] - data[j]) != flag_right:  # 出现趋势的改变，退出染色
                    break
                if j in type2:  # 搜到同类，退出染色
                    break
                if j in type1:  # 搜到另一类，染色
                    n_type2.append(j)
    for i in type1:
        if i not in type2:
            n_type1.append(i)
    return n_type1, n_type2

# 对单参数进行sst
def get_sst_score(values):
    for i in range(0, len(values)):
        values[i] = float(values[i])
    model = banpei.SST(w=25)
    results = model.detect(values)
    return results

# 对所有参数sst后进行mds，然后聚类
def multi_sst_mds(params, df):
    print("正在进行参数间相关性分析，请稍后")
    # 将所有可检测的参数对应的sst分数放入lst中
    lst = []
    for param in params:
        data = df[param].values.tolist()[3:]
        data = [float(t) for t in data]
        # 补齐部分nan,解决频率不一致问题
        for i in range(0, len(data)):
            # 向后寻找非nan数来补齐
            if data[i] != data[i]:
                for j in range(i + 1, len(data)):
                    if data[j] == data[j]:
                        data[i] = data[j]
                        break
            # 向前寻找非nan数来补齐
            if data[i] != data[i]:
                for j in range(i - 1, -1, -1):
                    if data[j] == data[j]:
                        data[i] = data[j]
                        break
        data = get_sst_score(data)
        lst.append(data)
    mds = manifold.MDS(n_components=2)
    param_mds = mds.fit_transform(lst)
    # k, sse = elbow_best_k_2d(param_mds)
    labels = KMeans(n_clusters=len(v_params) // 5).fit_predict(lst).tolist()
    # labels = KMeans(n_clusters=k).fit_predict(param_mds).tolist()
    # print(k)
    # print(len(v_params) // 4)
    # print(labels)
    # plt.scatter(np.array(param_mds).T[0], np.array(param_mds).T[1], c=labels)
    # plt.show()
    return labels

# sst方法下的获取最相关的参数
def get_most_related_sst(v_params, param_name, sst_mds_labels):
    most_related_params = []
    param_name_idx = v_params.index(param_name)
    target_label = sst_mds_labels[param_name_idx]
    for i in range(len(v_params)):
        if sst_mds_labels[i] == target_label:
            if v_params[i] == param_name:
                continue
            most_related_params.append(v_params[i])
    return most_related_params

# 分析该参数的完整性
def analyse_completeness(df, param_name):
    # 完整率：以有效数据的最小下标差作为频率，遍历所有下标，统计出缺失数据的下标及数量n，然后返回1-n/len(data)
    data = df[param_name].values.tolist()[3:]
    # tmp = []
    # for i in range(0, len(data)):
    #     tmp.append(float(data[i]))
    # data = tmp
    idx = []
    for i in range(len(data)):
        if data[i] == data[i]:
            idx.append(i)
    min_sep = 10000000
    for i in range(1, len(idx)):
        min_sep = min(min_sep, idx[i] - idx[i - 1])
    err_nums = 0
    for i in range(1, len(idx)):
        err_nums += (idx[i] - idx[i - 1] - 1) // min_sep
    return 1 - err_nums / len(data), min_sep

# 用mds做参数评分融合
def param_points_mix(param_points):
    mds = manifold.MDS(n_components=1)
    # pca = PCA(n_components=1)
    # kpca = KernelPCA(n_components=1)
    return mds.fit_transform(param_points)

# 设置dataframe保存为excel最合适的列宽
def to_excel_auto_column_weight(df: pd.DataFrame, writer: ExcelWriter, sheet_name):
    """DataFrame保存为excel并自动设置列宽"""
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    #  计算表头的字符宽度
    column_widths = (
        df.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
    )
    #  计算每列的最大字符宽度
    max_widths = (
        df.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
    )
    # 计算整体最大宽度
    widths = np.max([column_widths, max_widths], axis=0)
    # 设置列宽
    worksheet = writer.sheets[sheet_name]
    for i, width in enumerate(widths, 1):
        # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2


def reset_col(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename,sheet).fillna('-')
        df.loc[len(df)]=list(df.columns)                        #把标题行附加到最后一行
        for col in df.columns:
            index=list(df.columns).index(col)                   #列序号
            letter=get_column_letter(index+1)                   #列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()   #获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=collen*1.2+4     #也就是列宽为最大长度*1.2 可以自己调整
        for row_cells in wb.get_sheet_by_name(sheet):
            for cell in row_cells:
                cell.alignment = Alignment(horizontal='center')
    wb.save(filename)


if __name__ == '__main__':
    filename_set = "data/qar/148965-20160126.csv"
    root_path = filename_set[0:filename_set.rindex("/") + 1]
    csv_name = filename_set[filename_set.rindex("/") + 1:filename_set.rindex(".")]
    if not os.path.exists(root_path + csv_name + '_anomaly_detection_figs'):
        os.mkdir(root_path + csv_name + '_anomaly_detection_figs')
    print("请输入文件地址（最好为绝对路径）：", end="")
    filename = input()
    selected_param_name = "N11"
    df = read_data(filename_set)
    # 统计飞行阶段
    flight_phase_num = {}
    flight_phases = df['FLIGHT_PHASE'].values.tolist()[3:]
    for flight_phase in flight_phases:
        if flight_phase in flight_phase_num:
            flight_phase_num[flight_phase] += 1
        else:
            flight_phase_num[flight_phase] = 0

    params = extract_params(df)
    v_params = valid_params(params, df)
    sst_mds_labels = multi_sst_mds(v_params, df)
    # df_corr, df_p = get_corrcoef(v_params, df)
    result = pd.DataFrame()
    param_points = 0
    flying_points = 0
    # print(v_params)
    # print(df_corr)
    # print(df_p)
    # print(most_related_params)
    # print("该qar文件中存在以下可进行异常检测的参数，请选择一个进行单参数分析，请输入括号前的字符串，括号内的为详细解释：")
    # for param in v_params:
    #     detail = str(df[param][1])
    #     if detail == "nan":
    #         detail = "无详细描述"
    #     print(param + "（" + detail + "）")
    # param_name = input()

    for m in range(len(params)):
        param_name = params[m]
        print("正在处理参数" + param_name)
        plt.cla()
        completeness, freq = analyse_completeness(df, param_name)
        print("参数：" + param_name + "的完整率为：" + str(completeness))
        if param_name not in v_params:
            print("参数" + param_name + "不是数值型参数，无法进行异常检测")
            result[param_name] = [completeness, "can't detect", None]
            param_points += completeness
            continue
        # 使用sst+mds+聚类的方法找出相关点
        most_related_params = get_most_related_sst(v_params, param_name, sst_mds_labels)
        # param_name_idx = v_params.index(param_name)
        #     # target_label = sst_mds_labels[param_name_idx]
        #     # for i in range(len(v_params)):
        #     #     if sst_mds_labels[i] == target_label:
        #     #         if v_params[i] == param_name:
        #     #             continue
        #     #         most_related_params.append(v_params[i])

        # 使用相关分析的方法找出相关点
        # most_related_params = get_most_related_params(df_corr, df_p, param_name)

        values = df[param_name].values.tolist()[3:]
        tmp = []
        for i in range(0, len(values)):
            tmp.append(float(values[i]))
        tmp = [t for t in tmp if t == t]  # 删除nan
        values = tmp
        # print(len(values))
        x = range(0, len(values))
        # 将不同的飞行阶段用不同的颜色画在图上
        # flight_phase_y = {}
        # flight_phase_x = {}
        # for i in range(0, len(values)):
        #     if flight_phases[i] in flight_phase_y:
        #         flight_phase_y[flight_phases[i]].append(values[i])
        #     else:
        #         flight_phase_y[flight_phases[i]] = [values[i]]
        #     if flight_phases[i] in flight_phase_x:
        #         flight_phase_x[flight_phases[i]].append(i)
        #     else:
        #         flight_phase_x[flight_phases[i]] = [i]
        # flight_phases = list(set(flight_phases))
        # for flight_phase in flight_phases:
        #     y = flight_phase_y[flight_phase]
        #     x = flight_phase_x[flight_phase]
        #     ax1 = plt.subplot(1, 1, 1)
        #     plt.sca(ax1)
        #     plt.plot(x, y, linewidth=1)
        all_means = np.average(values)
        ax1 = plt.subplot(1, 1, 1)
        plt.sca(ax1)
        # for i in range(0, len(values)):
        #     plt.scatter(i, values[i], s=5, c=flight_phases_map[flight_phases[i]])
        plt.plot(x, values, linewidth=0.7)
        # plt.plot(x, [all_means] * len(values))

        # 窗口均值指标
        # values = df[param_name].values.tolist()[3:]
        # tmp = []
        # for i in range(0, len(values)):
        #     tmp.append(float(values[i]))
        # # 补齐部分nan,解决频率不一致问题
        # for i in range(0, len(tmp)):
        #     # 向后寻找非nan数来补齐
        #     if tmp[i] != tmp[i]:
        #         for j in range(i + 1, len(tmp)):
        #             if tmp[j] == tmp[j]:
        #                 tmp[i] = tmp[j]
        #                 break
        #     # 向前寻找非nan数来补齐
        #     if tmp[i] != tmp[i]:
        #         for j in range(i - 1, -1, -1):
        #             if tmp[j] == tmp[j]:
        #                 tmp[i] = tmp[j]
        #                 break
        # values = tmp
        # means = static_window_mean(values, 30)  #静态窗口
        means = dynamic_window_mean(values, flight_phases, flight_phase_num, freq)  #动态窗口
        # ax2 = plt.subplot(5, 1, 2)
        # plt.sca(ax2)
        # plt.plot(x, means)
        # 肘部法
        # k, sse = elbow_best_k(means)
        # labels = DBSCAN(eps=0.118, min_samples=10).fit_predict(np.array(anomaly_score).reshape(-1, 1)).tolist()
        labels1 = KMeans(n_clusters=2).fit_predict(np.array(means).reshape(-1, 1)).tolist()
        # plt.scatter(x, means, s=5, c=labels1)
        # ax3 = plt.subplot(5, 1, 3)
        # plt.sca(ax3)
        # plt.plot(range(2, 10), sse)
        # plt.scatter(k, sse[k - 2])

        # 左右斜率指标
        values = df[param_name].values.tolist()[3:]
        tmp = []
        for i in range(0, len(values)):
            tmp.append(float(values[i]))
        tmp = [t for t in tmp if t == t]  # 删除nan
        values = tmp
        slopes = point_slope(values)
        # ax4 = plt.subplot(5, 1, 4)
        # plt.sca(ax4)
        # plt.plot(x, slopes)
        # labels = DBSCAN(eps=0.118, min_samples=10).fit_predict(np.array(anomaly_score).reshape(-1, 1)).tolist()
        k, sse = elbow_best_k(slopes)
        labels2 = KMeans(n_clusters=k).fit_predict(np.array(slopes).reshape(-1, 1)).tolist()
        # plt.scatter(x, slopes, s=5, c=labels2)
        # ax5 = plt.subplot(5, 1, 5)
        # plt.sca(ax5)
        # plt.plot(range(2, 10), sse)
        # plt.scatter(k, sse[k - 2])

        # sst
        # model = banpei.SST(w=50)
        # values = df[param_name].values.tolist()[3:]
        # tmp = []
        # for i in range(0, len(values)):
        #     tmp.append(float(values[i]))
        # tmp = [t for t in tmp if t == t]  # 删除nan
        # values = tmp
        # sst_scores = model.detect(values)
        # labels3 = KMeans(n_clusters=2).fit_predict(np.array(sst_scores).reshape(-1, 1)).tolist()

        # pyod
        # detector_list = [KNN(n_neighbors=15), LOF(n_neighbors=20),
        #                  LOF(n_neighbors=25), LOF(n_neighbors=35),
        #                  COPOD(), IForest(n_estimators=100),
        #                  IForest(n_estimators=200)]

        # decide the number of parallel process, and the combination method
        # then clf can be used as any outlier detection model
        clf = LOF(30)
        clf.fit(np.array(values).reshape(-1, 1))
        labels4 = clf.labels_

        # 异常罕见原则，正常点占大多数
        set_labels = list(set(labels1))
        max_label = set_labels[0]
        max_cnt = -1
        for label in set_labels:
            if labels1.count(label) > max_cnt:
                max_cnt = labels1.count(label)
                max_label = label
        anomaly_points1 = []
        for i in range(0, len(labels1)):
            if labels1[i] != max_label:
                anomaly_points1.append(i)

        set_labels = list(set(labels2))
        max_label = set_labels[0]
        max_cnt = -1
        for label in set_labels:
            if labels2.count(label) > max_cnt:
                max_cnt = labels2.count(label)
                max_label = label
        anomaly_points2 = []
        for i in range(0, len(labels2)):
            if labels2[i] != max_label:
                anomaly_points2.append(i)

        # set_labels = list(set(labels3))
        # max_label = set_labels[0]
        # max_cnt = -1
        # for label in set_labels:
        #     if labels3.count(label) > max_cnt:
        #         max_cnt = labels3.count(label)
        #         max_label = label
        # anomaly_points3 = []
        # for i in range(0, len(labels3)):
        #     if labels3[i] != max_label:
        #         anomaly_points3.append(i)

        # 两种指标同时异常的才是最终的异常点
        anomaly_points = []
        for i in range(0, len(values)):
            # if i in anomaly_points1 and i in anomaly_points2:
            #     anomaly_points.append(i)
            if labels4[i] == 1:
                anomaly_points.append(i)
        # plt.plot(x, [means_avg] * len(means))
        # 多元线性回归得到两种不同的异常
        type1, type2 = multi_regression(df, param_name, most_related_params, anomaly_points)
        # 时间分析，染色法，修正多元回归得到的结果
        type1, type2 = time_analyse_cause(type1, type2, values)
        correctness = 1 - len(type1) / len(values)
        flying_points += 1 - len(type2) / (len(values) - len(type1))
        print("参数：" + param_name + "的准确率为：" + str(correctness))

        param_points += completeness * 0.6 + correctness * 0.4
        plt.sca(ax1)
        for p in type1:
            plt.scatter(p, values[p], s=1, c='g')
        for p in type2:
            plt.scatter(p, values[p], s=1, c='r')
        plt.savefig(root_path + csv_name + '_anomaly_detection_figs/' + param_name + '.jpg')
        result[param_name] = [completeness, correctness, '=HYPERLINK("%s", "%s figure")' % (csv_name + '_anomaly_detection_figs/' + param_name + '.jpg', param_name)]
    print(param_points)
    result["final_score"] = [flying_points / len(params), param_points / len(params), None]
    result.index = ["完整性（0~1，最后一列：驾驶评分）", "准确性（0~1，最后一列：质量评分）", "异常检测图像"]
    # with pd.ExcelWriter(root_path + csv_name + '_qar_report.xlsx') as writer:
    #     to_excel_auto_column_weight(result, writer, 'test')
    result.to_excel(root_path + csv_name + '_qar_report.xlsx')
    reset_col(root_path + csv_name + '_qar_report.xlsx')
    print("该qar文件已分析完毕，请查看qar_report.csv获取完整性和准确性的数值，进入anomaly_detection_figs查看每个参数的异常情况")

